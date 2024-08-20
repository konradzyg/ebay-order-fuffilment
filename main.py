import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

def remove_columns_and_highlight_orders(input_csv, output_excel, columns_to_remove):
    # Read the CSV file with the header in the second row (index 1)
    df = pd.read_csv(input_csv, header=1)
    
    df.drop(columns=columns_to_remove, inplace=True)
    
    order_numbers = input("Enter the order numbers to cancel, separated by spaces: ").split()
    
    # Save to an Excel file
    df.to_excel(output_excel, index=False)
    
    # Load the Excel file
    workbook = load_workbook(output_excel)
    sheet = workbook.active

    # Adjust column widths to fit the content
    for column in sheet.columns:
        max_length = 0
        column_letter = column[0].column_letter 
        for cell in column:
            try:
                max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        adjusted_width = max_length + 2
        sheet.column_dimensions[column_letter].width = adjusted_width
        
    # Find and highlight the rows in Excel
    for idx, row in df.iterrows():
        if str(row["Order Number"]) in order_numbers:
            for cell in sheet[idx+2]:  # +2 because Excel is 1-indexed and we skip the header row
                cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    
    # Save the Excel file
    workbook.save(output_excel)

if __name__ == "__main__":
    input_csv = 'ebay.csv'

    output_excel = 'output.xlsx'  

    # Columns you want to remove
    columns_to_remove = ['Sales Record Number', 'Buyer Name', 'Buyer Email', 'Buyer Note', 'Buyer Address 1', 'Buyer Address 2', 'Buyer City', 'Buyer State', 'Buyer Zip', 'Buyer Country', 'Buyer Tax Identifier Name', 'Buyer Tax Identifier Value', 'Ship To Phone', 'Ship To Address 1', 'Ship To Address 2', 'Ship To City', 'Ship To State', 'Ship To Zip', 'Ship To Country', 'Item Number', 'Item Title', 'Custom Label', 'Sold Via Promoted Listings', 'Quantity', 'Sold For', 'Shipping And Handling', 'Item Location', 'Item Zip Code', 'Item Country', 'eBay Collect And Remit Tax Rate', 'eBay Collect And Remit Tax Type', 'eBay Reference Name', 'eBay Reference Value', 'Tax Status', 'Seller Collected Tax', 'eBay Collected Tax', 'Electronic Waste Recycling Fee', 'Mattress Recycling Fee', 'Battery Recycling Fee', 'White Goods Disposal Tax', 'Tire Recycling Fee', 'Additional Fee', 'Lumber Fee', 'Prepaid Wireless Fee', 'Road Improvement And Food Delivery Fee', 'eBay Collected Charges', 'Total Price', 'eBay Collected Tax and Fees Included in Total', 'Payment Method', 'Sale Date', 'Paid On Date', 'Ship By Date', 'Minimum Estimated Delivery Date', 'Maximum Estimated Delivery Date', 'Shipped On Date', 'Feedback Left', 'Feedback Received', 'My Item Note', 'PayPal Transaction ID', 'Shipping Service', 'Tracking Number', 'Transaction ID', 'Variation Details', 'Global Shipping Program', 'Global Shipping Reference ID', 'Click And Collect', 'Click And Collect Reference Number', 'eBay Plus', 'Authenticity Verification Program', 'Authenticity Verification Status', 'Authenticity Verification Outcome Reason', 'PSA Vault Program', 'Vault Fulfillment Type', 'eBay Fulfillment Program', 'Tax City', 'Tax State', 'Tax Zip', 'Tax Country', 'eBay International Shipping']  # Replace with your columns

    remove_columns_and_highlight_orders(input_csv, output_excel, columns_to_remove)